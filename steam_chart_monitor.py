#!/usr/bin/env python3
"""
Steam 인기 차트 자동 추적기
매일 실행하여 상위 50개 게임의 지표를 Excel에 누적 저장합니다.

추적 지표:
  - 동시 접속자 수 (CCU)
  - 2주 플레이어 수 (판매 순위 대용)
  - 긍정 리뷰 비율 (%)
  - 현재 가격 및 할인율

롱런 기준: 28일(4주) 이상 상위 50위 유지
"""

import requests
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import date
import time
import os

# ── 설정 ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH   = os.path.join(SCRIPT_DIR, "steam_chart_tracker.xlsx")
TOP_N        = 50    # 상위 50개 게임 추적
LONGRUN_DAYS = 28    # 4주(28일) 이상 = 롱런 게임


# ── 데이터 수집 ───────────────────────────────────────────────────────────────

def fetch_steamspy_top100():
    """SteamSpy API: 최근 2주 상위 100개 게임"""
    url = "https://steamspy.com/api.php?request=top100in2weeks"
    r = requests.get(url, timeout=30)
    r.raise_for_status()
    data = r.json()

    games = []
    for rank, (appid, info) in enumerate(data.items(), 1):
        pos = info.get("positive", 0) or 0
        neg = info.get("negative", 0) or 0
        total = pos + neg
        review_pct = round(pos / total * 100, 1) if total > 0 else 0

        games.append({
            "rank":           rank,
            "appid":          int(appid),
            "name":           info.get("name", ""),
            "ccu":            info.get("ccu", 0) or 0,
            "players_2weeks": info.get("players_2weeks", 0) or 0,
            "review_score_pct": review_pct,
            "total_reviews":  total,
            "positive_reviews": pos,
        })

    return games[:TOP_N]


def fetch_steam_price(appid):
    """Steam Store API: 가격 및 할인율"""
    url = (
        f"https://store.steampowered.com/api/appdetails/"
        f"?appids={appid}&cc=kr&filters=price_overview"
    )
    try:
        r = requests.get(url, timeout=10)
        d = r.json()
        app = d.get(str(appid), {})
        if app.get("success") and app.get("data"):
            p = app["data"].get("price_overview", {})
            return {
                "price_krw":          p.get("final", 0) // 100,
                "discount_pct":       p.get("discount_percent", 0),
                "original_price_krw": p.get("initial", 0) // 100,
            }
    except Exception:
        pass
    return {"price_krw": None, "discount_pct": 0, "original_price_krw": None}


def collect_today_data():
    """오늘의 차트 전체 수집"""
    print("▶ SteamSpy 상위 50 게임 수집 중...")
    games = fetch_steamspy_top100()

    today_str = date.today().isoformat()
    rows = []
    for i, g in enumerate(games, 1):
        print(f"  [{i:2d}/{len(games)}] {g['name'][:40]}")
        price = fetch_steam_price(g["appid"])
        time.sleep(0.4)

        rows.append({
            "date":               today_str,
            "rank":               g["rank"],
            "appid":              g["appid"],
            "name":               g["name"],
            "ccu":                g["ccu"],
            "players_2weeks":     g["players_2weeks"],
            "review_score_pct":   g["review_score_pct"],
            "total_reviews":      g["total_reviews"],
            "positive_reviews":   g["positive_reviews"],
            "price_krw":          price["price_krw"],
            "discount_pct":       price["discount_pct"],
            "original_price_krw": price["original_price_krw"],
        })

    return pd.DataFrame(rows)


# ── 분석 ──────────────────────────────────────────────────────────────────────

def analyze_longrun(df):
    """4주(28일) 이상 상위 50위를 유지한 게임 집계"""
    if df.empty:
        return pd.DataFrame()

    df = df.copy()
    df["date"] = pd.to_datetime(df["date"])

    stats = df.groupby(["appid", "name"]).agg(
        days_in_top        = ("date",             "nunique"),
        avg_rank           = ("rank",             "mean"),
        best_rank          = ("rank",             "min"),
        avg_ccu            = ("ccu",              "mean"),
        latest_ccu         = ("ccu",              "last"),
        avg_review_score   = ("review_score_pct", "mean"),
        latest_review_score= ("review_score_pct", "last"),
        total_reviews      = ("total_reviews",    "last"),
        latest_price       = ("price_krw",        "last"),
        max_discount       = ("discount_pct",     "max"),
        first_seen         = ("date",             "min"),
        last_seen          = ("date",             "max"),
    ).reset_index()

    longrun = stats[stats["days_in_top"] >= LONGRUN_DAYS].copy()
    longrun.sort_values("days_in_top", ascending=False, inplace=True)

    longrun["avg_rank"]         = longrun["avg_rank"].round(1)
    longrun["avg_ccu"]          = longrun["avg_ccu"].round(0).astype(int)
    longrun["avg_review_score"] = longrun["avg_review_score"].round(1)
    longrun["first_seen"]       = longrun["first_seen"].dt.strftime("%Y-%m-%d")
    longrun["last_seen"]        = longrun["last_seen"].dt.strftime("%Y-%m-%d")

    return longrun


# ── Excel 출력 ────────────────────────────────────────────────────────────────

HDR_FILL = PatternFill("solid", start_color="1F4E79")   # 진한 남색
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL = PatternFill("solid", start_color="EBF3FB")   # 연한 파랑
DIS_FILL = PatternFill("solid", start_color="C6EFCE")   # 연한 초록 (할인)
LRN_FILL = PatternFill("solid", start_color="FFFACD")   # 연한 황금 (롱런)
CENTER   = Alignment(horizontal="center", vertical="center")


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_excel(all_df, today_df, longrun_df):
    wb = Workbook()

    # ── 시트 1: 일별 스냅샷 ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "일별 스냅샷"

    COLS = {
        "date":               "날짜",
        "rank":               "순위",
        "appid":              "AppID",
        "name":               "게임명",
        "ccu":                "동시접속자",
        "players_2weeks":     "2주 플레이어",
        "review_score_pct":   "긍정리뷰(%)",
        "total_reviews":      "총 리뷰수",
        "positive_reviews":   "긍정리뷰수",
        "price_krw":          "가격(₩)",
        "discount_pct":       "할인율(%)",
        "original_price_krw": "정가(₩)",
    }
    ws1.append(list(COLS.values()))
    _style_header(ws1, 1, len(COLS))

    keys = list(COLS.keys())
    for ri, row in enumerate(all_df.itertuples(index=False), 2):
        for ci, k in enumerate(keys, 1):
            ws1.cell(ri, ci, value=getattr(row, k, None))
        if ri % 2 == 0:
            for ci in range(1, len(keys) + 1):
                ws1.cell(ri, ci).fill = ALT_FILL

    _set_col_widths(ws1, [12, 5, 10, 36, 12, 14, 12, 12, 12, 10, 10, 10])
    ws1.freeze_panes = "A2"

    # ── 시트 2: 오늘의 차트 ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("오늘의 차트")
    title2 = f"Steam 인기 차트 — {date.today().isoformat()}"
    ws2["A1"] = title2
    ws2["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws2.append([])

    T_COLS = ["순위", "게임명", "동시접속자", "2주 플레이어", "긍정리뷰(%)", "총 리뷰수", "가격(₩)", "할인율(%)"]
    T_KEYS = ["rank", "name", "ccu", "players_2weeks", "review_score_pct", "total_reviews", "price_krw", "discount_pct"]
    ws2.append(T_COLS)
    _style_header(ws2, 3, len(T_COLS))

    for ri, row in enumerate(today_df.itertuples(index=False), 4):
        for ci, k in enumerate(T_KEYS, 1):
            ws2.cell(ri, ci, value=getattr(row, k, None))
        disc = getattr(row, "discount_pct", 0) or 0
        base_fill = DIS_FILL if disc > 0 else (ALT_FILL if ri % 2 == 0 else None)
        if base_fill:
            for ci in range(1, len(T_KEYS) + 1):
                ws2.cell(ri, ci).fill = base_fill

    _set_col_widths(ws2, [5, 36, 12, 14, 12, 12, 10, 10])
    ws2.freeze_panes = "A4"

    # ── 시트 3: 롱런 게임 분석 ──────────────────────────────────────────────
    ws3 = wb.create_sheet("롱런 게임 분석")
    ws3["A1"] = f"4주(28일) 이상 상위 50위 유지 게임 — 기준일: {date.today().isoformat()}"
    ws3["A1"].font = Font(bold=True, size=13, color="B8860B")
    ws3.append([])

    L_COLS = {
        "name":               "게임명",
        "days_in_top":        "유지 일수",
        "avg_rank":           "평균 순위",
        "best_rank":          "최고 순위",
        "avg_ccu":            "평균 동접",
        "latest_ccu":         "최근 동접",
        "avg_review_score":   "평균 긍정리뷰(%)",
        "latest_review_score":"최근 긍정리뷰(%)",
        "total_reviews":      "총 리뷰수",
        "latest_price":       "현재 가격(₩)",
        "max_discount":       "최대 할인율(%)",
        "first_seen":         "첫 관측일",
        "last_seen":          "최근 관측일",
    }

    ws3.append(list(L_COLS.values()))
    _style_header(ws3, 3, len(L_COLS))

    if not longrun_df.empty:
        for ri, row in enumerate(longrun_df.itertuples(index=False), 4):
            for ci, k in enumerate(L_COLS.keys(), 1):
                ws3.cell(ri, ci, value=getattr(row, k, None))
            for ci in range(1, len(L_COLS) + 1):
                ws3.cell(ri, ci).fill = LRN_FILL
    else:
        ws3["A4"] = (
            f"아직 {LONGRUN_DAYS}일 분량의 데이터가 쌓이지 않았습니다. "
            "매일 자동 수집이 진행되면 이 시트가 채워집니다."
        )
        ws3["A4"].font = Font(italic=True, color="888888")

    _set_col_widths(ws3, [36, 10, 10, 10, 12, 12, 16, 16, 12, 14, 14, 14, 14])
    ws3.freeze_panes = "A4"

    wb.save(EXCEL_PATH)
    print(f"✔ Excel 저장: {EXCEL_PATH}")


# ── 메인 ──────────────────────────────────────────────────────────────────────

def load_existing(path):
    if not os.path.exists(path):
        return pd.DataFrame()
    try:
        df = pd.read_excel(path, sheet_name="일별 스냅샷")
        df["date"] = pd.to_datetime(df["date"]).dt.date.astype(str)
        return df
    except Exception as e:
        print(f"기존 파일 로드 실패 ({e}), 새로 시작합니다.")
        return pd.DataFrame()


def main():
    today_str = date.today().isoformat()
    print(f"\n{'='*55}")
    print(f"  Steam 차트 추적기  |  {today_str}")
    print(f"{'='*55}")

    today_df   = collect_today_data()
    existing   = load_existing(EXCEL_PATH)

    # 오늘 데이터 중복 방지 후 병합
    if not existing.empty:
        existing = existing[existing["date"] != today_str]
    all_df = pd.concat([existing, today_df], ignore_index=True) if not existing.empty else today_df

    longrun_df = analyze_longrun(all_df.copy())
    print(f"\n▶ 롱런 게임 ({LONGRUN_DAYS}일+): {len(longrun_df)}개")

    build_excel(all_df, today_df, longrun_df)
    print("  완료!\n")


if __name__ == "__main__":
    main()
